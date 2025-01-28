from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import ElementClickInterceptedException, ElementClickInterceptedException, TimeoutException
from navigation import switch_to_iframe
from navigation import navigate_to_fecha_gen, navigate_to_carga_archivo_simple
from config import fecha
import time

def procesar_pendientes(driver, pendientes, summary_steps):
    """
    Procesa cada tabla pendiente ejecutando la función ejecutar_carga para cada una.
    """
    if not pendientes:
        print("No hay tablas pendientes para procesar.")
        return
    
    for i, id_tabla in enumerate(pendientes):
        try:
            print(f"Procesando tabla {i + 1}/{len(pendientes)}: {id_tabla}")
            # Ejecutar la función para procesar la tabla
            ejecutar_carga(driver, id_tabla)
            print(f"Tabla {id_tabla} procesada exitosamente.")
            summary_steps.append(f"→ Tabla {id_tabla} procesada exitosamente.")
        except Exception as e:
            # Captura cualquier error pero continúa con el siguiente pendiente
            print(f"Error al procesar la tabla {id_tabla}: {e}")
            continue  # Aseguramos que pase al siguiente elemento
    print("Todos los pendientes han sido procesados.")

#-------------------------------------------------------

def ejecutar_carga(driver, id_tabla):
    
    try:
        driver.switch_to.default_content()
        switch_to_iframe(driver, "e1menuAppIframe")

        # Localizar la tabla por su ID
        tabla = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, id_tabla))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tabla)
        print(f"Tabla {id_tabla} localizada y visible.")
        
        # Buscar el elemento dentro de la tabla y darle clic
        try:
            elemento_clic = tabla.find_element(
                By.XPATH, ".//td[@colindex='-2']//a/img[@title='Sin anexos']"
            )
            ActionChains(driver).move_to_element(elemento_clic).click().perform()
            print(f"Check en {id_tabla} HECHO.")
        except Exception as e:
            print(f"Error al localizar o hacer clic en el elemento dentro de la tabla {id_tabla}: {e}")
            return
        time.sleep(2)
        # Localizar y hacer doble clic en el botón "Ejecutar Carga"
        try:
            ejecutar_carga_boton = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.ID, "C0_28"))
            )
            ActionChains(driver).move_to_element(ejecutar_carga_boton).double_click().perform()
            print(f"Botón 'Ejecutar Carga' clickeado para la tabla {id_tabla}.")
        except Exception as e:
            print(f"Error al hacer clic en el botón 'Ejecutar Carga': {e}")
            return
        
        time.sleep(4)

        # Clic en el botón "Cancelar" dentro del iframe
        try:
            driver.switch_to.default_content()
            switch_to_iframe(driver, "e1menuAppIframe")
            herramientas_element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='WebLabel' and @title='Herramientas (Ctrl+Alt+T)']")))
            herramientas_element.click()

            buscar_cancel = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "hc_Cancel")))
            buscar_cancel.click()
            time.sleep(4)
            print("Clic en el botón 'Cancelar' realizado.")
        except Exception as e:
            print(f"Error al hacer clic en el botón 'Cancelar': {e}")
            return
        
        time.sleep(5)

        try:
            driver.switch_to.default_content()
            oracle_logo = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.ID, "oracleImage"))
            )
            oracle_logo.click()
            print("Clic en el logotipo de Oracle realizado para volver al inicio.")
        except Exception as e:
            print(f"Error al hacer clic en el logotipo de Oracle: {e}")
            return
        
        # Llamar a la función navigate_to_carga_archivo_simple
        navigate_to_carga_archivo_simple(driver)
        print(f"Usar función navigate_to_carga_archivo_simple en {id_tabla}")
        navigate_to_fecha_gen(driver, fecha)
        print(f"Usar función navigate_to_fecha_gen")
        print("------ Repeat ------")

    except Exception as e:
        print(f"Error en ejecutar_carga: {e}")


#-------------------------------------------------------

import openpyxl
import time
from openpyxl.utils import get_column_letter

def abrir_formato_registro_excel_escribir_batch_carga(driver):
    """
    Abrir el archivo 07.  FFN014-V1-Formato Registro BATCH-DICIEMBRE y escribir el batch de carga en la celda y hoja correspondiente.
    """
    # Ruta del archivo Excel
    ruta_excel = r"D:\OneDrive - Grupo EPM\Descargas\ResumenesIF\07.  FFN014-V1-Formato Registro BATCH-DICIEMBRE.xlsx"
    
    # Número de lote
    num_lote = 28954

    try:
        # Calcula el número de fila basado en la fecha
        dia = int(str(fecha)[-2:])  # Extrae los últimos dos dígitos y conviértelos a entero
        fila_destino = dia + 7  # Suma 7 al día

        # Nombre de la hoja donde escribir
        nombre_hoja = "1-FACTURACIÓN"

        # Abre el archivo Excel
        wb = openpyxl.load_workbook(ruta_excel)

        # Selecciona la hoja especificada
        if nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
        else:
            print(f"La hoja '{nombre_hoja}' no existe en el archivo.")
            return

        # Escribe el número de lote en la celda correspondiente (Columna B y fila calculada)
        columna = "B"
        celda_destino = f"{columna}{fila_destino}"
        hoja[celda_destino] = num_lote
        print(f"Número de lote {num_lote} escrito en la celda {celda_destino} de la hoja '{nombre_hoja}'.")

        # Guarda y cierra el archivo
        wb.save(ruta_excel)
        wb.close()
        print("Archivo Excel actualizado y guardado con éxito.")

        # Espera para visualizar el resultado (simulación visual)
        time.sleep(3)

    except Exception as e:
        print(f"Ocurrió un error al procesar el archivo Excel: {e}")


