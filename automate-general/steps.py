import time
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import pdfplumber  # Asegúrate de instalar esta biblioteca: pip install pdfplumber

# Configuración inicial
summary_steps = []
path = r"C:\Users\vgaleanc\Escritorio\chromedriver-win64\chromedriver.exe"
service = Service(executable_path=path)
driver = webdriver.Chrome(service=service)
summary_steps.append("Navegador Chrome abierto.")

# Abre el archivo Excel
excel_path = r"D:\OneDrive - Grupo EPM\Descargas\process.xlsx"
wb = load_workbook(excel_path)
ws = wb.active
summary_steps.append(f"Archivo Excel abierto desde '{excel_path}'.")

# Configura el Excel
ws["H5"] = "Total"
ws["H5"].font = Font(bold=True)
ws["I5"].number_format = "$ #.##0;-$ #.##0"
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws["I5"].border = border
summary_steps.append("Configuración inicial del Excel completada.")

# Abre la página de la Federación de Cafeteros
driver.execute_script("window.open('https://federaciondecafeteros.org/wp/publicaciones/', '_blank');")
summary_steps.append("Enlace de la Federación de Cafeteros abierto en una nueva pestaña.")
all_tabs = driver.window_handles
driver.switch_to.window(all_tabs[-1])

# Simula una interacción o tiempo de espera en la página de la Federación de Cafeteros
time.sleep(5)
summary_steps.append("Interacción con la página de Federación de Cafeteros completada.")

# Cambia al convertidor decimal-binario
driver.execute_script("window.open('https://www.prepostseo.com/tool/es/decimal-to-binary-converter', '_blank');")
summary_steps.append("Enlace del convertidor abierto en una nueva pestaña.")
all_tabs = driver.window_handles
driver.switch_to.window(all_tabs[-1])

# Captura el precio del café desde un PDF
try:
    wait = WebDriverWait(driver, 20)
    time.sleep(3)
    pdf_link = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "span.detail a")))
    pdf_link.click()
    summary_steps.append("Enlace al PDF del precio abierto.")

    # Cambia a la pestaña del PDF
    all_tabs = driver.window_handles
    driver.switch_to.window(all_tabs[-1])

    # Descargar el PDF
    pdf_path = r"D:\OneDrive - Grupo EPM\Descargas\precio_cafe.pdf"
    with open(pdf_path, "wb") as file:
        file.write(driver.find_element(By.TAG_NAME, 'body').screenshot_as_png())

    # Leer y procesar el PDF
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()

    precio_pasilla = None
    punto = None
    for line in text.splitlines():
        if "Precio total de pasilla contenida en el pergamino" in line:
            precio_pasilla = line.split()[-1]
        elif "Precio por punto producido" in line:
            punto = line.split()[-1].replace("COP", "").strip()

    if not punto:  # Si no se encuentra, usar valor por defecto
        punto = "800"
        summary_steps.append("Valor por punto no encontrado en el PDF. Usando valor por defecto: 800 COP.")

    summary_steps.append(f"Precio pasilla encontrado: {precio_pasilla}, punto capturado: {punto}.")
except Exception as e:
    summary_steps.append(f"Error al capturar el precio: {e}.")
    precio_pasilla = "Desconocido"
    punto = "800"

# Escribir resultados en el Excel
ws["A2"] = "El precio del café es:"
ws.column_dimensions['A'].width = 25
ws["B2"] = precio_pasilla
summary_steps.append("Escrito 'El precio del café es:' y el precio capturado en el Excel.")

# Convertir punto a binario
driver.switch_to.window(all_tabs[-1])
textarea = driver.find_element(By.ID, "x")
textarea.send_keys(punto)
summary_steps.append("Número del punto escrito en el convertidor.")

convert_button = driver.find_element(By.ID, "preloader")
convert_button.click()
summary_steps.append("Botón de convertir presionado.")

time.sleep(5)
puntobinario = driver.find_element(By.ID, "y").get_attribute("value")
summary_steps.append(f"Resultado del convertidor capturado: {puntobinario}.")

# Escribir resultados adicionales en el Excel
ws["A3"] = "PUNTO"
ws["A3"].font = Font(bold=True)
ws["B3"] = punto
ws["C3"] = "BINARIO"
ws["C3"].font = Font(bold=True)
ws["D3"] = puntobinario
summary_steps.append("Datos escritos en el Excel: punto y binario convertido.")

# Guardar Excel
wb.save(excel_path)
summary_steps.append(f"Archivo Excel guardado en '{excel_path}'.")

# Crear resumen de pasos
summary_path = r"D:\OneDrive - Grupo EPM\Descargas\RESUMEN.txt"
summary_steps.append(f"Resumen de pasos guardado en '{summary_path}'.")
summary_steps.append("Navegador cerrado.")
with open(summary_path, "w") as file:
    file.write("\n".join(summary_steps))

# Cerrar el navegador
driver.quit()

print("Automatización completada. Resumen guardado en archivo TXT.")
